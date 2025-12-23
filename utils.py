#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# TODO:
#       - Clean up this file
# ----------------------------------------------------------------------------
import os
from datetime import datetime
import decimal

from PyQt5 import QtWidgets, QtCore         # , QtGui
import qtawesome as qta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from gui.h_confirm_dialog import Ui_Dialog


# ---- Global Var ---- #
THEMES = {
    "light": {
        "stylesheet": "light_theme.qss",
        "colors": {
            "NEW_COLOR": "#1dd1a1",
            "MENU_COLOR": "#34495e",
            "ICON_COLOR": "black",
            "EDIT_COLOR": "#e67e22",
            "TRASH_COLOR": "#e74c3c",
            "BLUE_COLOR": "#3498db",
            "SAVE_COLOR": "#17c0eb",
            "SKYPE_COLOR": "#00AFF0",
        }
    },
    "dark": {
        "stylesheet": "dark_theme.qss",
        "colors": {
            "NEW_COLOR": "#1dd1a1",
            "MENU_COLOR": "#ecf0f1",
            "ICON_COLOR": "white",
            "EDIT_COLOR": "#d35400",
            "TRASH_COLOR": "#c0392b",
            "BLUE_COLOR": "#2980b9",
            "SAVE_COLOR": "#17c0eb",
            "SKYPE_COLOR": "#00AFF0",

        }
    }
}

NEW_COLOR = "#1dd1a1"
# NEW_COLOR = "#228447"
MENU_COLOR = "#DDDDDD"
SAVE_COLOR = '#17c0eb'
BLUE_COLOR = '#4074a3'
TRASH_COLOR = '#f77861'
EDIT_COLOR = '#FF6600'
WHITE_COLOR = "#FFFFFF"
ICON_COLOR = "#ececec"
SKYPE_COLOR = "#00AFF0"

Error_COLOR = "#f77861"
Success_COLOR = "#44e37b"

# Icons
EMPLOYES_HEADERS = ['ID', 'Nom', 'Phone', 'P. Travaille', 'Salaire', 'Date Embauche', 'Observation']
OPERATIONS_HEADERS = ['ID', 'Date', "Operation", 'Employé', 'Montant', 'Motif']
OPERATIONS_SUM_HEADERS = ["Employé", "T. Prime", "T. Retenu", "T. Avance"]

CLIENTS_HEADERS = ["ID", "Nom", "Crédit", "Telephone", "Commune", "Observation"]
CREDITS_HEADERS = ['ID', 'Date', 'Client', 'Par', 'Montant Total', 'Versement', 'Reste', 'Statut']
PAYMENTS_HEADERS = ['ID', 'Date', 'Client', 'Montant', 'Récupérateur']

CHARGE_HEADERS = ["ID", "Date", "Effectué par", "Montant", "Motif"]
VERSEMENT_HEADERS = ['ID', 'Date', 'Montant', 'Récupérateur']

# commune list
COMMUNES_LIST = [
    "TOUS",
    "CHIFALO", "KHMISTI", "BOUHAROUN", "BIRARD", "DAMOUS", "LARHAT",
    "BOUISMAIL", "SIDI-MOUSSA", "HAMDANIA", "BELDJ", "CHAIBA",
    "FOUKA", "BENHENNI", "AIN-LAHDJER", "SIDI-GHILES", "HADJRET-ENNOS", "DOUAOUDA",
    "DOUAOUDA-MARINE", "FOUKA-MARINE", "KOLEA", "TIPAZA", "NADOR", "CHERCHEL",
    "COMMUNEL", "CHAIG", "MESSELMOUN", "GHOURAYA", "HADJOUT",
]

MONTHS_FR = {
    "01": "JANVIER", "02": "FÉVRIER", "03": "MARS", "04": "AVRIL",
    "05": "MAI", "06": "JUIN", "07": "JUILLET", "08": "AOÛT",
    "09": "SEPTEMBRE", "10": "OCTOBRE", "11": "NOVEMBRE", "12": "DECEMBRE"
}


class ThemeManager:
    def __init__(self, app, themes, default="dark"):
        self.app = app
        self.themes = themes
        self.current = default
        self.apply(default)

    def apply(self, theme_name):
        theme = self.themes[theme_name]
        with open(f"./gui/{theme["stylesheet"]}", "r") as f:
            self.app.setStyleSheet(f.read())
        self.current = theme_name

    def color(self, role):
        return self.themes[self.current]["colors"][role]

    def icon(self, name, role):
        return qta.icon(name, color=self.color(role))


def format_money(value) -> str:
    """
    Format a number as money with comma as thousands separator and two decimals.
    Example: 12000 → '12,000.00'
    """
    try:
        # return "{:,.2f}".format(float(value))     # easy way
        # Frensh style with space
        value = float(value)
        parts = "{:,.2f}".format(value).split('.')
        integer_part = parts[0].replace(',', ' ')  # Replace comma with space
        decimal_part = parts[1]
        return f"{integer_part},{decimal_part}"
    except (ValueError, TypeError):
        return str(value)


def format_to_decimal(value):
    """
    Convert a string value to a Decimal, replacing spaces and commas.
    """
    # print(value)
    try:
        # Replace spaces and commas, then convert to Decimal
        value = value.replace(' ', '').replace(',', '.')
        return {'success': True, 'value': decimal.Decimal(value)}
    except (ValueError, TypeError, decimal.InvalidOperation):
        return {'success': False, 'error': 'Entrez un nombre valide.'}


def is_date(value: str, fmt="%Y-%m-%d") -> bool:
    """
    Docstring for is_date
    :type value: str
    :param fmt: Description
    :return: Description
    :rtype: bool
    """
    try:
        datetime.strptime(value, fmt)
        return True
    except ValueError:
        return False


# ==================
# == UI -- Functions
# ==================
# === 1) Setup callbacks/signals/menus (run once at startup) ===
def setup_main_callbacks(root):
    # --- Button callbacks (no icons here!)
    # root.ui.closeAppBtn.clicked.connect(root.close),
    # root.ui.minimizeAppBtn.clicked.connect(root.showMinimized),
    # root.ui.maximizeRestoreAppBtn.clicked.connect(root.toggle_maximize_restore),
    root.ui.toggleMenuButton.clicked.connect(root.on_toggle_menu),
    root.ui.extraCloseColumnBtn.clicked.connect(lambda: root.toggle_left_box(close=True)),
    root.ui.toggleThemeBtn.clicked.connect(root.toggle_theme),
    root.ui.buttonCloseMsgsFrame.clicked.connect(lambda: root.close_msgs_frame(close=True)),

    buttons = [
        # == Credit Page ==
        (root.ui.buttonCreditPage, lambda: root.goto_page("credit")),
        (root.ui.buttonNewCredit, root.ui_create_credit),
        (root.ui.buttonSaveCredit, root.save_new_credit),
        # == Credit Actions Edit/Versement/Delete
        (root.ui.buttonDeleteCredit, root.delete_credit),
        # ============================================================================================
        # == Payments Page ==
        (root.ui.buttonPaymentsPage, lambda: root.goto_page("payments")),
        (root.ui.buttonPaiementsEtatJournalier, root.paiements_etat_journalier),
        # versement for a specific credit
        (root.ui.buttonCreditVersement, root.credit_list_versement),

        # == Versement Small Table Page ==
        # ================================
        (root.ui.buttonDeletePayment, root.delete_payment),
        (root.ui.buttonCreditAddVersement, root.ui_add_versement),
        (root.ui.buttonSaveVersement, root.save_new_versement),
        (root.ui.buttonRegleCredit, root.regle_credit),
        (root.ui.buttonDeleteVersement, root.delete_versement),
        # ==================================================================================================
        # == Clients Page ==
        # ==================
        (
            # ph.users-three-thin
            # ph.users-three-light
            # ph.users-three
            root.ui.buttonClientsPage,
            lambda: root.goto_page('client')
        ),
        (root.ui.buttonNewClient, lambda: root.ui_create_persone('client')),
        (root.ui.buttonClientSituation, root.client_situation_report),
        (root.ui.buttonClientNewCredit, lambda: root.ui_create_credit(client=True)),
        (root.ui.buttonClientCreditList, root.client_credit_list),
        (root.ui.buttonDeleteClient, root.delete_client),
        # ================================================================================================
        # == Employes Page ==
        # ==================
        (root.ui.buttonEmployesPage, lambda: root.goto_page('employe')),
        (root.ui.buttonNewEmploye, lambda: root.ui_create_persone('employe')),
        # button save new both (EMPLOYE & CLIENTS)
        (root.ui.buttonSaveNewPerson, root.save_new_persone),
        (root.ui.buttonDeleteEmploye, root.delete_employe),

        # == Accompte Employee ==
        (
            # fa6s.sack-dollar
            # fa5s.file-invoice-dollar
            root.ui.buttonAccomptePage,
            lambda: root.goto_page('operations', from_btn=True)
        ),
        (root.ui.buttonEmployeNewAvance, lambda: root.ui_employe_opration('avance')),
        (root.ui.buttonEmployeNewPrime, lambda: root.ui_employe_opration('prime')),
        (root.ui.buttonEmployeNewRetenu, lambda: root.ui_employe_opration('retenu')),
        (root.ui.buttonCalculateSalaire, lambda: root.calculate_salaire(from_btn=True)),
        (root.ui.buttonExportAccomptDetails, root.export_accomptes_details),

        # Save Operation for Employees
        (root.ui.buttonEmployeSaveOperation, root.save_new_operation),
        # refresh to all
        (root.ui.buttonRefreshAccompteTable, lambda: root.display_accomptes(rows=None, headers_type="all")),
        (root.ui.buttonDeleteAccompte, root.delete_accompte),
        # ==================================================================================================
        # == Charge Page ==
        # ==================
        (root.ui.buttonChargePage, lambda: root.goto_page('charge', from_btn=True)),
        (root.ui.buttonNewCharge, root.ui_create_charge),
        (root.ui.buttonSaveCharge, root.insert_new_charge),
        (root.ui.buttonDeleteCharge, root.delete_charge),

        # ==========================================================
        # == Etats Page ==
        # ================
        (root.ui.buttonEtatPage, lambda: root.goto_page("etats")),
        (root.ui.buttonOpenEtatExcelFile, lambda: root.open_etat_excel_file(from_btn=True)),
        (root.ui.buttonEtatDetailJournee, root.etat_detail_journe),
    ]
    for button, callback in buttons:
        button.clicked.connect(callback)

    # === Tables ===
    tables = [
        (root.ui.clientsTableWidget, "client"),
        (root.ui.employesTableWidget, "employe"),
        (root.ui.creditTableWidget, "credit"),
        (root.ui.versementTableWidget, "payment"),
        (root.ui.accompteTableWidget, "operations"),
        (root.ui.chargeTableWidget, "charge"),
        (root.ui.paymentsTableWidget, "payments"),
    ]
    for table, page in tables:
        table.itemSelectionChanged.connect(lambda p=page: root.enable_buttons(p))

    table_edits = [
        (root.ui.employesTableWidget, root.edit_employe),
        (root.ui.accompteTableWidget, root.edit_accompte),
        (root.ui.clientsTableWidget, root.edit_client),
        (root.ui.creditTableWidget, root.edit_credit),
        (root.ui.chargeTableWidget, root.edit_charge),
        (root.ui.paymentsTableWidget, root.edit_payment),
    ]
    for table, callback in table_edits:
        table.editingFinished.connect(callback)

    root.ui.etatJournalierDetailsTableWidget.itemSelectionChanged.connect(root.etat_observation)
    # === ComboBoxes ===
    cbBoxes = [
        (root.ui.cbBoxCreditByStatus, lambda: root.filter_credit_by_status_commune(filter="status")),
        (root.ui.cbBoxCreditByCommune, lambda: root.filter_credit_by_status_commune(filter="commune")),
        (root.ui.cbBoxSalaireEmpMonth, lambda: root.calculate_salaire(from_btn=False)),
        (root.ui.cbBoxChargeByMonth, lambda: root.filter_charge()),
        (root.ui.cbBoxPaymentByMonth, root.payment_from_cbbox),
        # (root.ui.cbBoxEtatByMonth, root.display_etat_journalier),
    ]
    for cbBox, callback in cbBoxes:
        cbBox.currentIndexChanged.connect(callback)

    for cbBox in (
        root.ui.cbBoxEmployeOperationByName,
        root.ui.cbBoxEmployeOperationByType,
        root.ui.cbBoxEmployeOperationByDate
    ):
        cbBox.currentIndexChanged.connect(root.filter_accomptes)

    # === LineEdits ===
    root.ui.editSearchItem.textChanged.connect(root.search_item)
    root.ui.editSearchItem.returnPressed.connect(root.search_item)
    root.ui.buttonRefreshTableWidget.clicked.connect(root.refresh_table)

    root.ui.dateEditEtatJournee.dateChanged.connect(root.etat_detail_journe)

    # === Menus ===
    create_menu(
        root,
        root.ui.plusButtonShurtcut,
        "ph.plus",
        [
            ("Client", lambda: root.ui_create_persone('client'), "ph.plus"),
            ("Crédit", root.ui_create_credit, "ph.plus"),
            ("Employée", lambda: root.ui_create_persone('employee'), "ph.plus"),
            ("Charge", root.ui_create_charge, "ph.plus"),
        ],
        with_icons=True
    )
    create_menu(
        root,
        root.ui.buttonSettings,
        "fa6s.gear",
        [("Run Server", root.toggle_server, "mdi6.play-pause")],
        with_icons=True
    )
    create_menu(
        root,
        root.ui.buttonCreditActions,
        "ph.caret-down-bold",
        [("Exporté", lambda: root.excel_export_credits("credits"), "mdi6.microsoft-excel")],
        with_icons=True
    )
    create_menu(
        root,
        root.ui.buttonClientActions,
        "ph.caret-down-bold",
        [("Exporté", lambda: root.excel_export_credits(page="clients"), "mdi6.microsoft-excel")],
        with_icons=True
    )

    # === Context Menus ===
    employe_table_actions = [
        ('L. Accompte', qta.icon("fa6s.money-check-dollar", color=NEW_COLOR), root.accompte_by_employee),
        ('Calculer Salaire', qta.icon('mdi.calculator-variant', color=ICON_COLOR), lambda: root.calculate_salaire(from_btn=True)),
        ('separator', None, None),
        ('Supprimer', qta.icon('msc.trashcan', color=TRASH_COLOR), root.delete_employe),
    ]
    setup_table_context_menu(root.ui.employesTableWidget, employe_table_actions)

    client_table_actions = [
        ('N. Crédit', qta.icon('mdi6.cash-plus', color=NEW_COLOR), lambda: root.ui_create_credit(client=True)),
        ('L. Crédits', qta.icon('ph.list', color=ICON_COLOR), root.client_credit_list),
        ('L. Versement', qta.icon('fa6s.money-check-dollar', color=NEW_COLOR), root.client_versement_list),
        ('A. Versement', qta.icon('fa6s.hand-holding-dollar', color=NEW_COLOR), root.ui_add_versement),
        ('separator', None, None),
        ('Supprimer', qta.icon('msc.trashcan', color=TRASH_COLOR), root.delete_client),
    ]
    setup_table_context_menu(root.ui.clientsTableWidget, client_table_actions)

    credit_table_actions = [
        ('A. Versement', qta.icon('fa6s.hand-holding-dollar', color=NEW_COLOR), root.ui_add_versement),
        ('L. Versements', qta.icon('fa6s.money-check-dollar', color=NEW_COLOR), root.credit_list_versement),
        ('Régler', qta.icon('mdi6.cash-check', color=NEW_COLOR), root.regle_credit),
        ('separator', None, None),
        ('Supprimer', qta.icon('msc.trashcan', color=TRASH_COLOR), root.delete_credit),
    ]
    setup_table_context_menu(root.ui.creditTableWidget, credit_table_actions)

    charge_table_actions = [
        ('Modifier', qta.icon('ph.pencil-line-light', color=EDIT_COLOR), lambda: root.ui_create_charge(edit=True)),
        ('Supprimer', qta.icon('msc.trashcan', color=TRASH_COLOR), root.delete_charge),
    ]
    setup_table_context_menu(root.ui.chargeTableWidget, charge_table_actions)


# === 2) Refresh icons (call on startup + every theme toggle) ===
def refresh_main_icons(root, theme_manager: ThemeManager):
    PLUS_ICON = theme_manager.icon('ph.plus', "NEW_COLOR")
    CASH_PLUS_ICON = theme_manager.icon('mdi6.cash-plus', "NEW_COLOR")
    SAVE_ICON = theme_manager.icon('mdi.content-save', "BLUE_COLOR")
    TRASH_ICON = theme_manager.icon('msc.trashcan', "TRASH_COLOR")
    REFRESH_ICON = theme_manager.icon("mdi6.refresh", "ICON_COLOR")
    # EDIT_ICON       = theme_manager.icon('ph.pencil-line-light', "EDIT_COLOR")
    LIST_ICON = theme_manager.icon('ph.list', "ICON_COLOR")

    # --- Main Window Buttons
    # root.ui.closeAppBtn.setIcon(theme_manager.icon("ph.x", "MENU_COLOR"))
    # root.ui.minimizeAppBtn.setIcon(theme_manager.icon("mdi.window-minimize", "MENU_COLOR"))
    # root.ui.maximizeRestoreAppBtn.setIcon(theme_manager.icon("mdi.window-restore", "MENU_COLOR"))

    root.ui.toggleMenuButton.setIcon(theme_manager.icon("ri.menu-fold-fill", "ICON_COLOR"))
    root.ui.extraCloseColumnBtn.setIcon(theme_manager.icon("ph.x", "ICON_COLOR"))

    root.ui.buttonRefreshTableWidget.setIcon(REFRESH_ICON)

    # --- Credit Page
    root.ui.buttonCreditPage.setIcon(theme_manager.icon("ph.currency-circle-dollar", "MENU_COLOR"))
    root.ui.buttonCreditVersement.setIcon(qta.icon('fa6s.money-check-dollar', color=NEW_COLOR))
    root.ui.buttonNewCredit.setIcon(CASH_PLUS_ICON)
    root.ui.buttonSaveCredit.setIcon(SAVE_ICON)
    root.ui.buttonDeleteCredit.setIcon(TRASH_ICON)

    # --- Versements
    root.ui.buttonPaymentsPage.setIcon(theme_manager.icon("fa6s.hand-holding-dollar", "NEW_COLOR"))
    root.ui.buttonDeletePayment.setIcon(TRASH_ICON)

    root.ui.buttonCreditAddVersement.setIcon(theme_manager.icon('fa6s.hand-holding-dollar', "NEW_COLOR"))
    root.ui.buttonSaveVersement.setIcon(SAVE_ICON)
    root.ui.buttonRegleCredit.setIcon(qta.icon('mdi6.cash-check', color=NEW_COLOR))
    root.ui.buttonDeleteVersement.setIcon(TRASH_ICON)

    # --- Clients
    root.ui.buttonClientsPage.setIcon(theme_manager.icon('ph.users', "MENU_COLOR"))
    root.ui.buttonNewClient.setIcon(PLUS_ICON)
    root.ui.buttonClientNewCredit.setIcon(CASH_PLUS_ICON)
    root.ui.buttonClientCreditList.setIcon(LIST_ICON)
    root.ui.buttonDeleteClient.setIcon(TRASH_ICON)
    root.ui.buttonClientSituation.setIcon(theme_manager.icon('ri.file-excel-2-fill', "NEW_COLOR"))

    # --- Employés
    root.ui.buttonEmployesPage.setIcon(theme_manager.icon('mdi.account-hard-hat', "MENU_COLOR"))
    root.ui.buttonNewEmploye.setIcon(PLUS_ICON)
    root.ui.buttonSaveNewPerson.setIcon(SAVE_ICON)
    root.ui.buttonDeleteEmploye.setIcon(TRASH_ICON)

    root.ui.buttonAccomptePage.setIcon(theme_manager.icon('fa6s.file-invoice-dollar', "MENU_COLOR"))
    root.ui.buttonEmployeNewAvance.setIcon(CASH_PLUS_ICON)
    root.ui.buttonEmployeNewPrime.setIcon(CASH_PLUS_ICON)
    root.ui.buttonEmployeNewRetenu.setIcon(qta.icon('mdi6.cash-minus', color=TRASH_COLOR))
    root.ui.buttonCalculateSalaire.setIcon(theme_manager.icon('mdi.calculator-variant', "ICON_COLOR"))
    root.ui.buttonEmployeSaveOperation.setIcon(SAVE_ICON)
    root.ui.buttonRefreshAccompteTable.setIcon(REFRESH_ICON)
    root.ui.buttonDeleteAccompte.setIcon(TRASH_ICON)
    root.ui.buttonExportAccomptDetails.setIcon(theme_manager.icon('mdi6.microsoft-excel', "NEW_COLOR"))

    # --- Charges
    root.ui.buttonChargePage.setIcon(theme_manager.icon('mdi6.cash-minus', "EDIT_COLOR"))
    root.ui.buttonNewCharge.setIcon(PLUS_ICON)
    root.ui.buttonSaveCharge.setIcon(SAVE_ICON)
    root.ui.buttonDeleteCharge.setIcon(TRASH_ICON)

    # --- Etats
    # ph.chart-line-up
    root.ui.buttonEtatPage.setIcon(theme_manager.icon('ph.presentation-chart-thin', "ICON_COLOR"))
    root.ui.buttonOpenEtatExcelFile.setIcon(theme_manager.icon('mdi6.microsoft-excel', "NEW_COLOR"))

    # --- Standalone Icons
    root.ui.buttonIconSumPrime.setIcon(qta.icon('fa5s.comment-dollar', color=ICON_COLOR))
    root.ui.buttonIconSumAvance.setIcon(qta.icon('fa5s.comment-medical', color=ICON_COLOR))
    root.ui.buttonIconSumRetenu.setIcon(qta.icon('fa5s.dollar-sign', color=ICON_COLOR))
    root.ui.extraIconPlus.setIcon(qta.icon('ph.plus', color=SKYPE_COLOR))


def pagebuttons_stats(root):
    """
    Update page button states based on the current page.
    """
    ui = root.ui
    current_page = ui.stackedWidget.currentIndex()
    ui.buttonClientsPage.setChecked(current_page == 0)
    ui.buttonCreditPage.setChecked(current_page == 1)
    ui.buttonEmployesPage.setChecked(current_page == 2)
    ui.buttonAccomptePage.setChecked(current_page == 3)
    ui.buttonChargePage.setChecked(current_page == 4)
    ui.buttonPaymentsPage.setChecked(current_page == 5)
    ui.buttonEtatPage.setChecked(current_page == 6)


def clear_inputs(inputs: list) -> None:
    """
    This function clear inputs
    """
    for inp in inputs:
        inp.clear()


def remove_layout(widget):
    """
    Remove layout from a QWidget
    """
    layout = widget.layout()
    print(f"[+] Lyout: {layout}")
    if layout is None:
        return

    while layout.count():
        item = layout.takeAt(0)

        if item.widget():
            item.widget().deleteLater()
        elif item.layout():
            remove_layout_from_layout(item.layout())

    layout.deleteLater()


def remove_layout_from_layout(layout):
    """
    Remove a Layout from QLayout
    """
    while layout.count():
        item = layout.takeAt(0)
        if item.widget():
            item.widget().deleteLater()
        elif item.layout():
            remove_layout_from_layout(item.layout())


# ==================================
# == Table Widget Functions
# ==================================
def populate_table_widget(table: QtWidgets.QTableWidget, rows: list, headers: list, resize_to_content=False) -> None:
    """
    Populate a QTableWidget with rows and headers.

    :param table: The QTableWidget instance.
    :param rows: A list of rows where each row is a list or tuple of values.
    :param headers: A list of column headers.
    """
    table.clearContents()
    table.setColumnCount(len(headers))
    table.setRowCount(len(rows))
    table.setHorizontalHeaderLabels(headers)
    table.horizontalHeader().setVisible(True)
    table.setSortingEnabled(False)

    # These columns will be formatted as money
    money_headers = {
        'Salaire', 'Crédit', 'Montant Total', 'Versement', 'Reste',
        'Montant', 'Accomptes', 'Credits', 'Versements', 'Charges',
        "T. COMMANDE", "T.LOGICIEL", "VERSEMENT", "CHARGE", "DIFF"
    }
    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data):
            header = headers[col_idx].strip()
            if isinstance(value, (int, float)) and header in money_headers:
                text = format_money(value)
            else:
                text = str(value)
            item = QtWidgets.QTableWidgetItem(str(text))
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            table.setItem(row_idx, col_idx, item)

    table.horizontalHeader().setStretchLastSection(True)
    table.setSortingEnabled(True)
    if resize_to_content:
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)


def table_has_selection(table: QtWidgets.QTableWidget) -> bool:
    """
    Check if table has a selected rows
    :table: table widget name
    :return: True or False
    """
    if len(table.selectionModel().selectedRows()) > 0: return True
    else: return False


def get_column_value(table: QtWidgets.QTableWidget, row: int, column: int) -> str:
    """
    Get the value from a specific column of the selected row in a QTableWidget.

    :param table: The QTableWidget instance.
    :param column: The column index to retrieve the value from.
    :return: The value as a string.
    """
    return table.item(row, column).text()


def table_multi_selection(table: QtWidgets.QTableWidget) -> list:
    """
    This function return column(0) for a selection
    :table: QTableWidget
    :return: a list of ids.
    """
    selected_rows = set(index.row() for index in table.selectedIndexes())   # return index of selected row
    ids = list()
    if len(selected_rows) > 0:
        for row in selected_rows:
            item_id = table.item(row, 0).text()
            ids.append(item_id)
    return ids


def set_table_column_sizes(table_widget, *sizes):
    """
    Set the width of each column in a QTableWidget.

    Args:
        table_widget (QTableWidget): The table widget to modify.
        *sizes (int): Variable number of column widths. Each value sets the width for the corresponding column.
    """
    for col, size in enumerate(sizes):
        table_widget.setColumnWidth(col, size)


def setup_table_context_menu(table_widget: QtWidgets.QTableWidget, actions: list):
    """
    Adds a right-click context menu to the provided QTableWidget.
    :param table_widget: The QTableWidget instance.
    """
    table_widget.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
    table_widget.customContextMenuRequested.connect(
        lambda pos: show_table_context_menu(table_widget, actions, pos)
    )


def show_table_context_menu(table_widget: QtWidgets.QTableWidget, actions: list, pos: QtCore.QPoint):
    """
    Display a context menu at the given position.

    :param table_widget: The QTableWidget instance.
    :param pos: The position of the right-click.
    """
    index = table_widget.indexAt(pos)
    if not index.isValid():
        return

    menu = QtWidgets.QMenu(table_widget)
    for label, icon, callback in actions:
        if label == 'separator':
            menu.addSeparator()
            continue
        action = QtWidgets.QAction(icon, label, table_widget) if icon else QtWidgets.QAction(label, table_widget)
        action.triggered.connect(callback)
        menu.addAction(action)

    menu.exec_(table_widget.viewport().mapToGlobal(pos))


def export_tablewidget_to_excel(table: QtWidgets.QTableWidget, file_name: str = "export", title: str = "Export") -> str:
    """
    Export the contents of a QTableWidget to an Excel file with a date and title.
    File is saved inside 'excel_fichier' directory, auto-created if missing.

    :param table: The QTableWidget instance.
    :param file_name: Base file name (without extension).
    :param title: The title to display in the Excel file.
    :return: Full path of the saved file.
    """
    # ---- Ensure directory exists ----
    output_dir = "excel_fichier"
    os.makedirs(output_dir, exist_ok=True)

    # ---- Build filename with timestamp ----
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(output_dir, f"{file_name}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active

    # ---- Date in first line ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table.columnCount())
    ws["A1"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A1"].font = Font(bold=True, italic=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="right")

    # ---- Title in second line ----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=table.columnCount())
    ws["A2"] = title
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ---- Headers in third line ----
    headers = []
    for col in range(table.columnCount()):
        header_item = table.horizontalHeaderItem(col)
        headers.append(header_item.text() if header_item else f"Column {col + 1}")
    ws.append(headers)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---- Data rows ----
    for row in range(table.rowCount()):
        row_data = []
        for col in range(table.columnCount()):
            item = table.item(row, col)
            row_data.append(item.text() if item else "")
        ws.append(row_data)

    # ---- Auto column width ----
    for col in range(1, table.columnCount() + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # ---- Save file ----
    wb.save(filename)
    return f"✅ Exporté {table.rowCount()} ligne X {table.columnCount()} avec succès dans '{filename}'."


# == QComboBox
def populate_comboBox(combobox: QtWidgets.QComboBox, items: list):
    """
    Populate a QComboBox with a list of items.

    :param combobox: The QComboBox instance.
    :param items: A list of strings to populate the combobox.
    """
    combobox.blockSignals(True)
    combobox.clear()
    combobox.addItems(items)
    combobox.blockSignals(False)


def create_menu(root, menu_button, icon_name, actions, icon_color=NEW_COLOR, with_icons=False):
    """
    Create and attach a menu to a QPushButton.

    :param root: Parent widget.
    :param menu_button: QPushButton to attach the menu to.
    :param icon_name: Name of the icon to set on the button (QtAwesome).
    :param actions: List of tuples:
                    - (label, callback) if with_icons=False
                    - (label, callback, icon_path) if with_icons=True
    :param with_icons: Whether actions include icons.
    """
    # Set main button icon
    menu_button.setIcon(qta.icon(icon_name, color=icon_color))

    # Create the menu
    menu = QtWidgets.QMenu(root)

    for action in actions:
        if with_icons:
            label, callback, icon_path = action
            act = menu.addAction(qta.icon(icon_path, color=NEW_COLOR), label)
        else:
            label, callback = action
            act = menu.addAction(label)

        act.triggered.connect(callback)

    # Attach menu to button
    menu_button.setMenu(menu)


# == Excel functions
def export_salary_report_openpyxl(rows, file_path):
    """
    Create a nicely formatted Excel report from rows using openpyxl.
    All tables are placed in one sheet.

    Args:
        rows (list of tuples): Each tuple = (id, date, type, name, somme, motif)
        file_name (str): Name of the Excel file to save.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Group rows by 'name'
    grouped = {}
    for r in rows:
        _, date, type_, name, somme, motif = r
        grouped.setdefault(name, []).append((date, type_, somme, motif))

    row_cursor = 1
    header_font = Font(bold=True)
    # total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # light yellow
    for name, entries in grouped.items():
        salaire_base = 0

        # Person's name as a title
        ws.cell(row=row_cursor, column=1, value=name).font = Font(bold=True, size=14)
        row_cursor += 1

        # Table header
        headers = ["Date", "Type", "Somme", "Motif"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        row_cursor += 1

        # Counters
        total_avance = total_retenu = total_prime = 0

        # Rows
        for date, type_, somme, motif in entries:
            ws.cell(row=row_cursor, column=1, value=date)
            ws.cell(row=row_cursor, column=2, value=type_)
            ws.cell(row=row_cursor, column=3, value=somme)
            ws.cell(row=row_cursor, column=4, value=motif)

            if type_ == "avance":
                total_avance += somme
            elif type_ == "retenu":
                total_retenu += somme
            elif type_ == "prime":
                total_prime += somme

            row_cursor += 1

        # Final salary calculation
        salaire_final = salaire_base + total_prime - total_retenu - total_avance

        # Summary block
        ws.append([])
        ws.cell(row=row_cursor, column=2, value="Salaire de base").font = header_font
        ws.cell(row=row_cursor, column=3, value=salaire_base)
        row_cursor += 1

        ws.cell(row=row_cursor, column=2, value="Total Prime").font = header_font
        ws.cell(row=row_cursor, column=3, value=total_prime)
        row_cursor += 1

        ws.cell(row=row_cursor, column=2, value="Total Retenu").font = header_font
        ws.cell(row=row_cursor, column=3, value=total_retenu)
        row_cursor += 1

        ws.cell(row=row_cursor, column=2, value="Total Avance").font = header_font
        ws.cell(row=row_cursor, column=3, value=total_avance)
        row_cursor += 1

        final_cell = ws.cell(row=row_cursor, column=2, value="Salaire Final")
        final_cell.font = Font(bold=True, color="FFFFFF")
        final_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # green
        ws.cell(row=row_cursor, column=3, value=salaire_final).font = Font(bold=True)

        row_cursor += 2  # space before next table

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    return f"✅ Accompte Exporté avec succès dans '{file_path}'."


def export_situation_to_excel(rows, file_path):
    # TODO: fix this function
    """
    Export the client's situation (credits + versements) in a structured Excel table.
    Each credit and its versements are grouped with clear separation lines.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Versements"

    # Informations client
    ws.append(["Client", rows[0]['client']])
    ws.append(["Total Crédit", rows[0]['total_montant']])
    ws.append(["Reste", rows[0]['reste']])
    # Calculate total versement
    sum_versement = sum(r['montant'] for r in rows[1:])
    ws.append(["Total Versement", sum_versement])

    # --- Add this to format the numbers as money ---
    ws["B2"].number_format = '#,##0'   # Total Crédit
    ws["B3"].number_format = '#,##0'   # Reste
    ws["B4"].number_format = '#,##0'   # Total Versement

    # Empty lines
    ws.append([])
    ws.append([])

    # En-têtes du tableau
    ws.append(["Date de versement", "Montant", "Recupérateur"])

    # Lignes du tableau
    payments = rows[1:]
    for p in payments:
        ws.append([p['date_versement'], p['montant'], p['recuperateur']])

    # Définir le tableau Excel
    start_row = 7  # Ligne où commence le tableau (après infos client)
    end_row = start_row + len(payments)
    table_ref = f"A{start_row}:C{end_row}"

    table = Table(displayName="TableVersements", ref=table_ref)
    style = TableStyleInfo(name="TableStyleLight8", showRowStripes=True)
    table.tableStyleInfo = style

    ws.add_table(table)

    # Auto-size des colonnes
    for col in range(1, 4):
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            val = row[0].value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    table.showTotals = True
    table.totalsRowShown = True

    # Sauvegarde du fichier
    wb.save(file_path)
    return f"✔️ Fichier généré : {file_path}"


class ConfirmDialog(QtWidgets.QDialog):
    def __init__(self, title):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        # Remove title bar
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WidgetAttribute.WA_TranslucentBackground)
        self.ui.labelTitle.mouseMoveEvent = self.move_window  # to move window from the upBar

        self.ui.labelMessage.setText(title)
        self.ui.buttonConfirm.clicked.connect(self.accept)
        self.ui.buttonCancel.clicked.connect(self.reject)

    # -- Window UPBAR Controls --
    def mousePressEvent(self, event):
        self.clickPosition = event.globalPos()

    def move_window(self, e):
        """Move the window from upBar"""
        if not self.isMaximized():
            if e.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + e.globalPos() - self.clickPosition)
                self.clickPosition = e.globalPos()
                e.accept()


if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)

    dialog = ConfirmDialog("Confirm Action")
    dialog.show()
    sys.exit(app.exec_())
